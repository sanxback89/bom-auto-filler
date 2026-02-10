"""
Excel 템플릿 탐색 및 스타일 관련 헬퍼
- Master 라벨 위치 찾기
- BOM 헤더 행/열 매핑
- 행 스타일 복사, 행 용량 확보
- 열 너비 자동 조정
"""
from typing import Dict, Tuple
from copy import copy

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from utils import clean_text, normalize_header


# ----------------------------
# Master 필드 라벨 정의
# ----------------------------
MASTER_LABELS = {
    "design_number": ["Design Number"],
    "description": ["Description"],
    "bom_number": ["BOM Number"],
    "legacy_style_numbers": ["Legacy Style Numbers"],
    "hang_fold_instructions": ["Hang/Fold Instructions"],
}

BOM_REQUIRED_HEADERS = [
    "Product",
    "Material Name",
    "Supplier Article Number",
    "Usage",
    "Quality Details",
    "Supplier [Allocate]",
]


def find_master_value_cells(ws: Worksheet) -> Dict[str, Tuple[int, int]]:
    """
    Finds where to write master values (typically in col B next to labels in col A).
    Returns mapping field_key -> (row, col) to write the value.
    """
    write_cells: Dict[str, Tuple[int, int]] = {}

    for r in range(1, min(ws.max_row, 80) + 1):
        for c in range(1, min(ws.max_column, 15) + 1):
            cell_val = clean_text(ws.cell(r, c).value)
            if not cell_val:
                continue
            for key, labels in MASTER_LABELS.items():
                if any(normalize_header(cell_val) == normalize_header(lb) for lb in labels):
                    write_cells[key] = (r, c + 1)

    missing = [k for k in MASTER_LABELS.keys() if k not in write_cells]
    if missing:
        raise ValueError(f"템플릿에서 Master 라벨을 못 찾음: {missing}")
    return write_cells


def find_bom_header_row_and_cols(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    """
    Finds BOM header row containing the required headers.
    Returns (header_row, col_map) where col_map is normalized_header -> column_index.
    """
    target_norm = {normalize_header(h): h for h in BOM_REQUIRED_HEADERS}
    optional_norms = {normalize_header("Image"), normalize_header("Color"), normalize_header("Category")}

    for r in range(1, ws.max_row + 1):
        norms_in_row = []
        for c in range(1, ws.max_column + 1):
            nv = normalize_header(ws.cell(r, c).value)
            if nv:
                norms_in_row.append(nv)

        if "product" in norms_in_row and "materialname" in norms_in_row:
            col_map: Dict[str, int] = {}
            for c in range(1, ws.max_column + 1):
                nv = normalize_header(ws.cell(r, c).value)
                if nv in target_norm:
                    col_map[nv] = c

            for c in range(1, ws.max_column + 1):
                nv = normalize_header(ws.cell(r, c).value)
                if nv in optional_norms:
                    col_map[nv] = c

            missing = [normalize_header(h) for h in BOM_REQUIRED_HEADERS if normalize_header(h) not in col_map]
            if not missing:
                return r, col_map

    raise ValueError("템플릿에서 BOM Details 헤더 행(Product/Material Name/.../Supplier [Allocate])을 찾지 못했습니다.")


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, start_col: int, end_col: int, color_base_col: int):
    """Copy styles from src_row to dst_row between start_col~end_col."""
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    for c in range(start_col, end_col + 1):
        src_cell = ws.cell(src_row, c)
        if src_cell is None:
            src_cell = ws.cell(src_row, color_base_col)

        dst_cell = ws.cell(dst_row, c)
        dst_cell._style = copy(src_cell._style)
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def ensure_bom_rows_capacity(
    ws: Worksheet,
    start_row: int,
    needed_rows: int,
    bom_start_col: int,
    style_end_col: int,
    color_base_col: int,
):
    """Ensure there are at least `needed_rows` writable rows starting at start_row."""
    if needed_rows <= 0:
        return

    max_existing_row = start_row - 1
    for (r, c) in ws._cells.keys():
        if r >= start_row and bom_start_col <= c <= style_end_col:
            if r > max_existing_row:
                max_existing_row = r

    existing_capacity = max(0, max_existing_row - start_row + 1)

    if existing_capacity >= needed_rows:
        return

    to_add = needed_rows - existing_capacity
    insert_at = max_existing_row + 1

    ws.insert_rows(insert_at, amount=to_add)

    for i in range(to_add):
        dst = insert_at + i
        copy_row_style(
            ws=ws,
            src_row=start_row,
            dst_row=dst,
            start_col=bom_start_col,
            end_col=style_end_col,
            color_base_col=color_base_col,
        )


def adjust_column_widths(ws: Worksheet):
    """Automatically adjust column widths based on content length."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_value = str(cell.value)
                    if '\n' in cell_value:
                        cell_length = max(len(line) for line in cell_value.split('\n'))
                    else:
                        cell_length = len(cell_value)
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 100)
        adjusted_width = max(adjusted_width, 10)
        
        ws.column_dimensions[column_letter].width = adjusted_width
