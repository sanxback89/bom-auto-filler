"""
GUI 모듈 - tkinter 기반 사용자 인터페이스
"""
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from excel_writer import fill_template, fill_sheet, sanitize_sheet_name


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BOM PDF → Excel Template Auto Filler (Multi)")
        self.geometry("680x500")

        self.template_path = tk.StringVar()
        self.saved_template = None

        self._build_ui()

    def _build_ui(self):
        pad = 8

        # Template
        tk.Label(self, text="1) 엑셀 양식(.xlsx) - 한 번만 선택하면 재사용됩니다").grid(
            row=0, column=0, sticky="w", padx=pad, pady=(pad, 2)
        )
        tk.Entry(self, textvariable=self.template_path, width=65).grid(
            row=1, column=0, padx=pad, sticky="w"
        )
        tk.Button(self, text="찾기", command=self.browse_template, width=10).grid(
            row=1, column=1, padx=pad
        )

        # PDF (복수 선택)
        tk.Label(self, text="2) BOM PDF(.pdf) - 여러 파일 선택 가능, 선택 시 자동 실행됩니다").grid(
            row=2, column=0, sticky="w", padx=pad, pady=(pad, 2)
        )
        tk.Entry(self, text="", width=65, state="readonly").grid(
            row=3, column=0, padx=pad, sticky="w"
        )
        tk.Button(self, text="찾기", command=self.browse_pdfs_and_run, width=10).grid(
            row=3, column=1, padx=pad
        )

        # Progress bar
        progress_frame = tk.Frame(self)
        progress_frame.grid(row=4, column=0, columnspan=2, padx=pad, pady=(pad, 2), sticky="ew")

        self.progress = ttk.Progressbar(progress_frame, mode="determinate", length=550)
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.progress_label = tk.Label(progress_frame, text="", width=12, anchor="e")
        self.progress_label.pack(side=tk.RIGHT, padx=(6, 0))

        # Log
        tk.Label(self, text="로그").grid(row=5, column=0, sticky="w", padx=pad, pady=(pad, 2))

        scroll_frame = tk.Frame(self)
        scroll_frame.grid(row=6, column=0, columnspan=2, padx=pad, pady=(2, pad), sticky="nsew")

        scrollbar = tk.Scrollbar(scroll_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log = tk.Text(scroll_frame, height=15, width=80, yscrollcommand=scrollbar.set)
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log.yview)

        self.grid_rowconfigure(6, weight=1)
        self.grid_columnconfigure(0, weight=1)

    def _reset_progress(self):
        self.progress["value"] = 0
        self.progress_label.config(text="")
        self.update_idletasks()

    def _set_progress(self, current: int, total: int):
        pct = int(current / total * 100) if total else 0
        self.progress["value"] = pct
        self.progress_label.config(text=f"{current}/{total}")
        self.update_idletasks()

    def browse_template(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.template_path.set(path)
            self.saved_template = path
            self._log(f"✅ 템플릿 선택됨: {os.path.basename(path)}")
            self._log("   → 이 템플릿은 앞으로 계속 사용됩니다\n")

    def browse_pdfs_and_run(self):
        """복수 PDF 선택 후 자동으로 실행 - 하나의 파일, 시트별 분리"""
        if not self.saved_template or not os.path.exists(self.saved_template):
            self._log("⚠️  먼저 엑셀 양식을 선택해주세요!\n")
            messagebox.showwarning("템플릿 필요", "먼저 엑셀 양식 파일을 선택해주세요.")
            return

        paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if not paths:
            return

        total = len(paths)
        self._reset_progress()

        self._log("=" * 70)
        self._log(f"📋 선택된 PDF: {total}개")
        for i, p in enumerate(paths, 1):
            self._log(f"   {i}. {os.path.basename(p)}")
        self._log("=" * 70 + "\n")

        output_dir = os.path.dirname(paths[0])

        try:
            if total == 1:
                # 단일 PDF → 별도 파일
                pdf_path = paths[0]
                pdf_basename = os.path.splitext(os.path.basename(pdf_path))[0]
                output_path = os.path.join(output_dir, f"{pdf_basename}_filled.xlsx")

                self._log(f"📄 [1/1] 처리 중: {os.path.basename(pdf_path)}")
                self._set_progress(0, 1)
                saved = fill_template(self.saved_template, pdf_path, output_path)
                self._set_progress(1, 1)
                self._log(f"   ✅ 완료: {os.path.basename(saved)}")
            else:
                # 복수 PDF → 하나의 파일, 시트별 분리
                output_path = os.path.join(output_dir, "BOM_combined_filled.xlsx")

                wb = load_workbook(self.saved_template)
                original_sheet_names = list(wb.sheetnames)
                template_ws = wb.active

                sheet_names_used = set()
                success_count = 0
                fail_count = 0

                for idx, pdf_path in enumerate(paths, 1):
                    self._set_progress(idx - 1, total)
                    self._log(f"📄 [{idx}/{total}] 처리 중: {os.path.basename(pdf_path)}")

                    try:
                        new_ws = wb.copy_worksheet(template_ws)
                        design_number = fill_sheet(new_ws, pdf_path)

                        # 시트 이름 결정
                        name = design_number or os.path.splitext(os.path.basename(pdf_path))[0]
                        name = sanitize_sheet_name(name)

                        base_name = name
                        counter = 1
                        while name in sheet_names_used:
                            suffix = f"_{counter}"
                            name = sanitize_sheet_name(base_name[:31 - len(suffix)] + suffix)
                            counter += 1
                        sheet_names_used.add(name)
                        new_ws.title = name

                        self._log(f"   ✅ 완료 → 시트: {name}")
                        success_count += 1

                    except Exception as e:
                        self._log(f"   ❌ 실패: {str(e)}")
                        fail_count += 1

                    self._set_progress(idx, total)

                # 원본 템플릿 시트 모두 삭제
                for sn in original_sheet_names:
                    if sn in wb.sheetnames:
                        wb.remove(wb[sn])

                wb.save(output_path)

                if fail_count > 0:
                    self._log(f"\n   ⚠️ 성공: {success_count}개 / 실패: {fail_count}개")

            self._log("\n" + "=" * 70)
            self._log(f"📊 작업 완료!")
            self._log(f"   📁 저장 위치: {output_path}")
            self._log("=" * 70 + "\n")

            messagebox.showinfo("완료", f"작업 완료!\n\n저장: {os.path.basename(output_path)}")

        except Exception as e:
            self._log(f"   ❌ 실패: {str(e)}")
            self._log("=" * 70 + "\n")
            messagebox.showerror("오류", f"처리 중 오류 발생:\n\n{str(e)}")

    def _log(self, msg: str):
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.update_idletasks()
