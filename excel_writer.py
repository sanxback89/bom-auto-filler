"""
Excel 쓰기 모듈 - fill_template 메인 함수
PDF에서 파싱한 데이터를 엑셀 양식에 채워넣는 핵심 로직
"""
import os
import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border
from openpyxl.utils import get_column_letter

from models import group_rows_by_material
from pdf_parser import parse_master_from_pdf, extract_bom_rows_from_pdf
from image_handler import insert_design_image_into_sheet, insert_bom_row_image
from excel_template import (
    find_master_value_cells,
    find_bom_header_row_and_cols,
    ensure_bom_rows_capacity,
)


def _has_any_border(cell) -> bool:
    b = cell.border
    if b is None:
        return False
    return any(
        getattr(side, "style", None)
        for side in (b.left, b.right, b.top, b.bottom)
    )


def _fill_sheet(ws, pdf_path: str) -> str:
    """
    워크시트 하나에 PDF BOM 데이터를 채워넣는 핵심 로직.
    Returns: design_number (시트 이름용)
    """
    # 1) Find where to write master fields
    master_cells = find_master_value_cells(ws)

    # 2) Parse PDF master fields
    master = parse_master_from_pdf(pdf_path)

    # 3) Write master
    ws.cell(*master_cells["design_number"]).value = master.get("design_number", "")
    ws.cell(*master_cells["description"]).value = master.get("description", "")
    ws.cell(*master_cells["bom_number"]).value = master.get("bom_number", "")
    ws.cell(*master_cells["legacy_style_numbers"]).value = master.get("legacy_style_numbers", "")
    ws.cell(*master_cells["hang_fold_instructions"]).value = master.get("hang_fold_instructions", "")

    # 3.5) Insert Design Image
    try:
        insert_design_image_into_sheet(ws, pdf_path)
    except Exception:
        pass

    # 4) Find BOM header row + columns
    header_row, col_map = find_bom_header_row_and_cols(ws)
    start_row = header_row + 1

    c_category = col_map.get("category")
    c_product = col_map["product"]
    c_material = col_map["materialname"]
    c_supp_art = col_map["supplierarticlenumber"]
    c_usage = col_map["usage"]
    c_quality = col_map["qualitydetails"]
    c_supplier = col_map.get("supplierallocate") or col_map.get("supplier")
    c_image = col_map.get("image")

    if c_supplier is None:
        c_supplier = c_quality

    c_color_start = c_supplier + 1

    # 5) Parse BOM table rows + color headers
    raw_rows, color_headers = extract_bom_rows_from_pdf(pdf_path)
    grouped_rows = group_rows_by_material(raw_rows)

    # Insert subtitle rows when section starts.
    output_rows = []
    prev_cat_norm = ""
    subtitle_by_cat = {
        "packaging and labels": "Packaging and Labels",
    }
    for r in grouped_rows:
        cat_norm = (r.category or "").strip().lower()
        if cat_norm in subtitle_by_cat and prev_cat_norm != cat_norm:
            output_rows.append(("subtitle", subtitle_by_cat[cat_norm]))
        output_rows.append(("data", r))
        prev_cat_norm = cat_norm

    # 6) Determine color columns count
    num_color_cols = len(color_headers or [])
    style_start_col = c_category if c_category else c_product
    style_end_base = max(c_supplier, c_quality, c_image or 0)

    # Detect original template table end column from header/body border footprint.
    template_end_col = style_end_base
    blank_streak = 0
    scan_row_end = min(ws.max_row, start_row + 2)
    for cc in range(style_start_col, ws.max_column + 1):
        marked = False
        for rr in range(header_row, scan_row_end + 1):
            cell = ws.cell(rr, cc)
            if _has_any_border(cell) or bool(cell.value):
                marked = True
                break
        if marked:
            template_end_col = cc
            blank_streak = 0
        else:
            blank_streak += 1
            if cc > template_end_col and blank_streak >= 8:
                break

    original_color_capacity = max(0, template_end_col - c_color_start + 1)
    dynamic_end_col = c_color_start + max(num_color_cols - 1, 0)
    style_end_col = max(style_end_base, template_end_col, dynamic_end_col)

    # Capacity originally prepared in template (before inserting extra rows).
    max_existing_row = start_row - 1
    for (r, c) in ws._cells.keys():
        if r >= start_row and style_start_col <= c <= template_end_col and r > max_existing_row:
            max_existing_row = r
    original_capacity = max(0, max_existing_row - start_row + 1)

    # 6.5) Write color headers
    if num_color_cols > 0:
        base_header_cell = ws.cell(header_row, c_supplier)
        for j, htxt in enumerate(color_headers):
            cc = c_color_start + j
            cell = ws.cell(header_row, cc)
            if cell._style is None or cell._style == ws.cell(1, 1)._style:
                cell._style = copy(base_header_cell._style)
                cell.font = copy(base_header_cell.font)
                cell.border = copy(base_header_cell.border)
                cell.fill = copy(base_header_cell.fill)
                cell.number_format = base_header_cell.number_format
                cell.protection = copy(base_header_cell.protection)
                cell.alignment = copy(base_header_cell.alignment)

            cell.value = htxt
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # 7) Ensure enough rows
    ensure_bom_rows_capacity(
        ws=ws,
        start_row=start_row,
        needed_rows=len(output_rows),
        bom_start_col=style_start_col,
        style_end_col=style_end_col,
        color_base_col=c_supplier,
    )

    # If color columns exceed template slots, force expanded column width to 36.13
    # and copy border/style from last template color column when available.
    if num_color_cols > original_color_capacity:
        extra_start = c_color_start + original_color_capacity
        extra_end = c_color_start + num_color_cols - 1
        src_col_for_style = c_color_start + original_color_capacity - 1 if original_color_capacity > 0 else c_supplier
        style_row_end = max(start_row + max(original_capacity, len(output_rows)) - 1, header_row)
        for cc in range(extra_start, extra_end + 1):
            ws.column_dimensions[get_column_letter(cc)].width = 36.13
            for rr in range(header_row, style_row_end + 1):
                src = ws.cell(rr, src_col_for_style)
                dst = ws.cell(rr, cc)
                dst._style = copy(src._style)
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)

    # 8) Write BOM details rows
    for i, row_item in enumerate(output_rows):
        rr = start_row + i

        for cc in range(style_start_col, style_end_col + 1):
            ws.cell(rr, cc).value = None

        if row_item[0] == "subtitle":
            subtitle_col = c_category if c_category else c_product
            subtitle_cell = ws.cell(rr, subtitle_col)
            subtitle_cell.value = f"[ {row_item[1]} ]"
            subtitle_cell.alignment = Alignment(vertical="center", horizontal="left")
            subtitle_cell.font = copy(subtitle_cell.font)
            subtitle_cell.font = subtitle_cell.font.copy(bold=True)
            continue

        r = row_item[1]
        if c_category:
            ws.cell(rr, c_category).value = r.category
        ws.cell(rr, c_product).value = r.product
        ws.cell(rr, c_material).value = r.material_name
        ws.cell(rr, c_supp_art).value = r.supplier_article_number
        ws.cell(rr, c_usage).value = r.usage
        ws.cell(rr, c_quality).value = r.quality_details
        ws.cell(rr, c_supplier).value = r.supplier

        if c_image and getattr(r, "image_png", None):
            try:
                insert_bom_row_image(ws, rr, c_image, r.image_png)
            except Exception:
                pass

        for j, htxt in enumerate(color_headers):
            v = (r.colors or {}).get(htxt, "")
            if v:
                ws.cell(rr, c_color_start + j).value = v
            if getattr(r, "color_images", None) and htxt in r.color_images:
                try:
                    insert_bom_row_image(ws, rr, c_color_start + j, r.color_images[htxt])
                except Exception:
                    pass

    # 9) If parsed data is smaller than template capacity, clean remaining area.
    #    Keep layout size, but remove borders for empty rows.
    filled_rows = len(output_rows)
    if original_capacity > filled_rows:
        for rr in range(start_row + filled_rows, start_row + original_capacity):
            for cc in range(style_start_col, style_end_col + 1):
                cell = ws.cell(rr, cc)
                cell.value = None
                cell.border = Border()

    # 10) If parsed colors are fewer than template color slots, clear unused color area
    #     (same rows) to avoid bordered empty blocks.
    if num_color_cols < original_color_capacity:
        unused_start = c_color_start + num_color_cols
        unused_end = c_color_start + original_color_capacity - 1
        row_end = start_row + max(filled_rows, original_capacity) - 1
        for rr in range(header_row, row_end + 1):
            for cc in range(unused_start, unused_end + 1):
                cell = ws.cell(rr, cc)
                cell.value = None
                cell.border = Border()

    return master.get("design_number", "")


# public alias
fill_sheet = _fill_sheet


def sanitize_sheet_name(name: str) -> str:
    """Excel 시트 이름 규칙에 맞게 정리 (최대 31자, 특수문자 제거)"""
    for ch in ['\\', '/', '?', '*', '[', ']', ':']:
        name = name.replace(ch, '')
    name = name.strip()[:31]
    if not name:
        name = "Sheet"
    return name


def fill_template(
    template_path: str,
    pdf_path: str,
    output_path: str,
) -> str:
    """단일 PDF → 단일 Excel 파일 (기존 동작 유지)"""
    wb = load_workbook(template_path)
    ws = wb.active
    _fill_sheet(ws, pdf_path)
    wb.save(output_path)
    return output_path
