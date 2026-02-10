"""
ì´ë¯¸ì§€ ì¶”ì¶œ ë° Excel ì‚½ìž… ê´€ë ¨ í•¨ìˆ˜ë“¤
- Design Image: PDF ì²« íŽ˜ì´ì§€ ìŠ¤ì¼€ì¹˜ ì¶”ì¶œ â†’ Excel ì‚½ìž…
- BOM Row Image: Packaging/Graphic ì„¹ì…˜ ì´ë¯¸ì§€ ì¶”ì¶œ â†’ Excel ì…€ ì‚½ìž…
"""
import re
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import pdfplumber
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyxlImage

from utils import clean_text, normalize_header, clean_text_keep_newlines, format_color_header_text
from models import section_from_cell_text

try:
    import fitz as _fitz  # PyMuPDF – 렌더링 없이 임베디드 이미지 직접 추출
except ImportError:
    _fitz = None

from PIL import Image as PILImage

TARGET_BOM_IMAGE_WIDTH_CM = 1.88
_PX_PER_INCH = 96.0
_CM_PER_INCH = 2.54

# Full-page render cache (avoids re-rendering the same page for each cell)
_page_render_cache: Dict = {}


def _crop_cell_image(page, bbox, resolution=200):
    """
    Full-page render → pixel-level crop.
    page.crop(bbox).to_image() 방식은 인접 셀의 임베디드 이미지를
    정확히 분리하지 못하는 버그가 있어, 전체 페이지를 한 번 렌더링 후
    픽셀 좌표로 크롭하는 방식으로 대체.
    """
    global _page_render_cache
    cache_key = (id(page), resolution)

    if cache_key not in _page_render_cache:
        _page_render_cache.clear()
        _page_render_cache[cache_key] = page.to_image(resolution=resolution).original

    full_img = _page_render_cache[cache_key]

    x0, top, x1, bottom = bbox
    page_w = float(page.width)
    page_h = float(page.height)
    img_w, img_h = full_img.size

    scale_x = img_w / page_w
    scale_y = img_h / page_h

    px_x0 = max(0, int(x0 * scale_x))
    px_top = max(0, int(top * scale_y))
    px_x1 = min(img_w, int(x1 * scale_x))
    px_bottom = min(img_h, int(bottom * scale_y))

    if px_x1 <= px_x0 or px_bottom <= px_top:
        return None

    return full_img.crop((px_x0, px_top, px_x1, px_bottom))


# ----------------------------
# PyMuPDF 직접 이미지 추출 (플랫폼 독립)
# ----------------------------
_fitz_image_cache: Dict[Tuple[str, int], List[Tuple[Tuple[float, float, float, float], bytes]]] = {}


def _get_fitz_images_for_page(
    pdf_path: str, page_idx: int
) -> List[Tuple[Tuple[float, float, float, float], bytes]]:
    """
    PyMuPDF로 페이지의 모든 임베디드 이미지를 추출.
    렌더링 없이 PDF 내부의 원본 이미지 데이터를 직접 가져오므로
    OS/백엔드에 무관하게 동일한 결과를 보장.
    Returns: [((x0, y0, x1, y1), png_bytes), ...]
    """
    if _fitz is None:
        return []
    cache_key = (pdf_path, page_idx)
    if cache_key in _fitz_image_cache:
        return _fitz_image_cache[cache_key]

    results: List[Tuple[Tuple[float, float, float, float], bytes]] = []
    doc = None
    try:
        doc = _fitz.open(pdf_path)
        page = doc[page_idx]
        img_list = page.get_images(full=True)
        processed_xrefs: set = set()

        for img_info in img_list:
            xref = img_info[0]
            if xref in processed_xrefs:
                continue
            processed_xrefs.add(xref)
            try:
                rects = page.get_image_rects(xref)
                if not rects:
                    continue
                base = doc.extract_image(xref)
                if not base or not base.get("image"):
                    continue
                raw = base["image"]
                pil = PILImage.open(BytesIO(raw))
                if pil.mode == "CMYK":
                    pil = pil.convert("RGB")
                elif pil.mode not in ("RGB", "RGBA", "L"):
                    pil = pil.convert("RGB")
                buf = BytesIO()
                pil.save(buf, format="PNG")
                png = buf.getvalue()
                for rect in rects:
                    if rect.is_empty or rect.is_infinite:
                        continue
                    results.append(((rect.x0, rect.y0, rect.x1, rect.y1), png))
            except Exception:
                continue
    except Exception:
        pass
    finally:
        if doc:
            doc.close()

    _fitz_image_cache[cache_key] = results
    return results


def _find_fitz_image_for_bbox(
    pdf_path: str,
    page_idx: int,
    bbox: Tuple[float, float, float, float],
    min_overlap: float = 25.0,
) -> Optional[bytes]:
    """
    셀 bbox와 가장 많이 겹치는 임베디드 이미지를 PyMuPDF로 직접 추출.
    렌더링 기반이 아니므로 Windows/Linux 무관하게 올바른 이미지 반환.
    """
    images = _get_fitz_images_for_page(pdf_path, page_idx)
    if not images:
        return None

    x0, top, x1, bottom = bbox
    cell_area = max(1.0, (x1 - x0) * (bottom - top))

    best_score = 0.0
    best_png: Optional[bytes] = None

    for (ix0, iy0, ix1, iy1), png_bytes in images:
        ow = max(0.0, min(x1, ix1) - max(x0, ix0))
        oh = max(0.0, min(bottom, iy1) - max(top, iy0))
        overlap = ow * oh
        if overlap < min_overlap:
            continue
        img_area = max(1.0, (ix1 - ix0) * (iy1 - iy0))
        score = overlap / min(cell_area, img_area)
        if score > best_score:
            best_score = score
            best_png = png_bytes

    return best_png


# ----------------------------
# Pixel helpers
# ----------------------------
def _col_width_to_pixels(width: Optional[float]) -> int:
    w = width if (width is not None and width > 0) else 8.43
    return int(w * 7 + 5)


def _row_height_to_pixels(height: Optional[float]) -> int:
    h = height if (height is not None and height > 0) else 15.0
    return int(h * 4 / 3)


def _pixels_to_col_width(px: float) -> float:
    if px <= 0:
        return 8.43
    return max(1.0, (px - 5.0) / 7.0)


def _pixels_to_row_height_points(px: float) -> float:
    if px <= 0:
        return 15.0
    return max(1.0, px * 72.0 / _PX_PER_INCH)


def _cm_to_pixels(cm: float) -> int:
    return max(1, int(round((cm / _CM_PER_INCH) * _PX_PER_INCH)))


def _trim_pil_to_content(pil_img):
    """
    Trim large white margins so the sketch doesn't become a huge blank rectangle.
    Only trim LEFT/RIGHT to avoid cutting top/bottom of sketches.
    """
    try:
        w, h = pil_img.size
        if w <= 0 or h <= 0:
            return pil_img

        pad = max(2, int(min(w, h) * 0.01))
        inner = pil_img.crop((pad, pad, max(pad + 1, w - pad), max(pad + 1, h - pad)))

        gray = inner.convert("L")
        bw = gray.point(lambda p: 255 if p < 245 else 0)
        bbox = bw.getbbox()
        if bbox is None:
            return pil_img

        x0, y0, x1, y1 = bbox
        expand = max(6, int(min(inner.size) * 0.03))
        x0 = max(0, x0 - expand)
        x1 = min(inner.size[0], x1 + expand)

        return inner.crop((x0, 0, x1, inner.size[1]))
    except Exception:
        return pil_img


def _has_embedded_image_in_bbox(page, bbox: Tuple[float, float, float, float]) -> bool:
    """Check if PDF has an embedded image overlapping the cell bbox."""
    try:
        x0, top, x1, bottom = bbox
        for im in (page.images or []):
            ix0 = float(im.get("x0", 0))
            ix1 = float(im.get("x1", 0))
            itop = float(im.get("top", 0))
            ibot = float(im.get("bottom", 0))
            w = max(0.0, min(x1, ix1) - max(x0, ix0))
            h = max(0.0, min(bottom, ibot) - max(top, itop))
            if w * h > 25.0:
                return True
        return False
    except Exception:
        return False


def _is_blank(pil_img) -> bool:
    try:
        w, h = pil_img.size
        pad = max(2, int(min(w, h) * 0.03))
        if w > pad * 2 and h > pad * 2:
            pil_img = pil_img.crop((pad, pad, w - pad, h - pad))

        gray = pil_img.convert("L")
        px = gray.getdata()
        total = len(px) if px is not None else 0
        if total == 0:
            return True
        nonwhite = sum(1 for p in px if p < 245)
        ratio = nonwhite / total
        return ratio < 0.01
    except Exception:
        return True


# ----------------------------
# Design Image (ì²« íŽ˜ì´ì§€ ìŠ¤ì¼€ì¹˜)
# ----------------------------
def find_design_image_anchor_and_box(ws: Worksheet) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    """
    Find the 'Design Image' label cell and return:
      - anchor (row, col) where the image should be placed
      - box size (width_px, height_px) representing the target area to fit into
    """
    label_row = None
    label_col = None

    for r in range(1, min(ws.max_row, 80) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            v = clean_text(ws.cell(r, c).value)
            if normalize_header(v) == "designimage":
                label_row, label_col = r, c
                break
        if label_row is not None:
            break

    if label_row is None or label_col is None:
        raise ValueError("í…œí”Œë¦¿ì—ì„œ 'Design Image' ë¼ë²¨ì„ ì°¾ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.")

    anchor_row = label_row
    anchor_col = label_col + 1

    target_range = None
    for mr in ws.merged_cells.ranges:
        if mr.min_row == anchor_row and mr.min_col == anchor_col:
            target_range = mr
            break
    if target_range is None:
        for mr in ws.merged_cells.ranges:
            if (mr.min_row <= anchor_row <= mr.max_row) and (mr.min_col <= anchor_col <= mr.max_col):
                target_range = mr
                break

    if target_range is None:
        def has_any_border(r: int, c: int) -> bool:
            b = ws.cell(r, c).border
            return any(getattr(getattr(b, side), "style", None) for side in ["left", "right", "top", "bottom"])

        bom_details_row = None
        for rr in range(anchor_row, min(ws.max_row, anchor_row + 60) + 1):
            v = clean_text(ws.cell(rr, 1).value)
            if normalize_header(v) == "bomdetails":
                bom_details_row = rr
                break

        min_row = anchor_row
        max_row = (bom_details_row - 1) if (bom_details_row and bom_details_row > anchor_row) else min(anchor_row + 10, ws.max_row)

        min_col = anchor_col
        max_col = anchor_col
        while max_col < ws.max_column and has_any_border(min_row, max_col + 1):
            max_col += 1

        max_col = min(max_col, anchor_col + 8)
        max_row = min(max_row, anchor_row + 20)
    else:
        min_row, max_row = target_range.min_row, target_range.max_row
        min_col, max_col = target_range.min_col, target_range.max_col

    width_px = 0
    for c in range(min_col, max_col + 1):
        letter = get_column_letter(c)
        width_px += _col_width_to_pixels(ws.column_dimensions[letter].width)

    height_px = 0
    for r in range(min_row, max_row + 1):
        height_px += _row_height_to_pixels(ws.row_dimensions[r].height)

    return (min_row, min_col), (max(1, width_px), max(1, height_px))


def extract_design_image_from_pdf(pdf_path: str):
    """
    Extract the sketch image area from the first page of the PDF.
    Returns a PIL Image (or None if extraction fails).
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return None
            page = pdf.pages[0]
            # 0) Prefer extracting the embedded image itself
            try:
                page_images = page.images or []
                if page_images:
                    cand = []
                    for im in page_images:
                        x0 = float(im.get("x0", 0))
                        x1 = float(im.get("x1", 0))
                        top = float(im.get("top", 0))
                        bottom = float(im.get("bottom", 0))
                        area = max(0.0, (x1 - x0)) * max(0.0, (bottom - top))
                        if area < 5000:
                            continue
                        bonus = 1.2 if top < page.height * 0.6 else 1.0
                        cand.append((area * bonus, (x0, top, x1, bottom)))
                    if cand:
                        _, bbox = max(cand, key=lambda x: x[0])
                        x0, top, x1, bottom = bbox
                        pad = 2.0
                        bbox = (
                            max(0.0, x0 - pad),
                            max(0.0, top - pad),
                            min(page.width, x1 + pad),
                            min(page.height, bottom + pad),
                        )
                        im = _crop_cell_image(page, bbox, resolution=200)
                        if im is None:
                            raise ValueError("empty crop")
                        im = _trim_pil_to_content(im)
                        return im
            except Exception:
                pass

            # 1) Fallback: locate by header text then crop
            words = page.extract_words() or []

            y_bottom = None
            for i in range(len(words) - 1):
                if normalize_header(words[i].get("text")) == "design" and normalize_header(words[i + 1].get("text")) == "image":
                    y_bottom = float(words[i + 1].get("bottom"))
                    break
            if y_bottom is None:
                image_words = [w for w in words if normalize_header(w.get("text")) == "image"]
                image_words = [w for w in image_words if float(w.get("top", 0)) < page.height * 0.45]
                if image_words:
                    w0 = sorted(image_words, key=lambda w: float(w.get("top", 0)))[0]
                    y_bottom = float(w0.get("bottom"))

            next_headers = {"tech", "techpack", "components", "documents", "bom", "measurement", "disclaimer"}
            y_next = None
            if y_bottom is not None:
                for w in words:
                    t = normalize_header(w.get("text"))
                    if t in next_headers and float(w.get("top")) > y_bottom + 10:
                        y_next = float(w.get("top"))
                        break

            if y_bottom is None:
                top = page.height * 0.18
            else:
                top = max(0, min(page.height, y_bottom - 2))

            if y_next is None:
                bottom = min(page.height, page.height * 0.70)
            else:
                bottom = max(top + 10, min(page.height, y_next - 6))

            bbox = (0, top, page.width, bottom)
            im = _crop_cell_image(page, bbox, resolution=200)
            if im is None:
                return None
            im = _trim_pil_to_content(im)
            return im
    except Exception:
        return None


def insert_design_image_into_sheet(ws: Worksheet, pdf_path: str):
    """Insert the first-page Design Image into the template sheet."""
    pil_img = extract_design_image_from_pdf(pdf_path)
    if pil_img is None:
        return

    # Keep template layout unchanged: always anchor Design image at B6.
    ar, ac = 6, 2

    # Fit into merged area containing B6 when present; otherwise use B6 cell box.
    min_r, min_c, max_r, max_c = _get_merged_box(ws, ar, ac)
    box_w = 0
    for c in range(min_c, max_c + 1):
        letter = get_column_letter(c)
        box_w += _col_width_to_pixels(ws.column_dimensions[letter].width)
    box_h = 0
    for r in range(min_r, max_r + 1):
        box_h += _row_height_to_pixels(ws.row_dimensions[r].height)
    box_w = max(1, box_w)
    box_h = max(1, box_h)

    iw, ih = pil_img.size
    if iw <= 0 or ih <= 0:
        return
    scale = 0.98 * min(box_w / iw, box_h / ih)
    scale = min(scale, 1.0) if scale > 0 else 1.0
    target_w = int(iw * scale)
    target_h = int(ih * scale)

    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    img = OpenPyxlImage(buf)
    img.width = target_w
    img.height = target_h

    ws.add_image(img, ws.cell(ar, ac).coordinate)

    try:
        if hasattr(img, "anchor") and hasattr(img.anchor, "ext") and img.anchor.ext is not None:
            img.anchor.ext.cx = int(target_w * 9525)
            img.anchor.ext.cy = int(target_h * 9525)
    except Exception:
        pass


# ----------------------------
# BOM Row Image (ì…€ ë‚´ ì´ë¯¸ì§€ ì‚½ìž…)
# ----------------------------
def _get_merged_box(ws: Worksheet, row: int, col: int) -> Tuple[int, int, int, int]:
    for mr in ws.merged_cells.ranges:
        if (mr.min_row <= row <= mr.max_row) and (mr.min_col <= col <= mr.max_col):
            return mr.min_row, mr.min_col, mr.max_row, mr.max_col
    return row, col, row, col


def insert_bom_row_image(ws: Worksheet, row: int, col: int, image_png: bytes,
                         scale_factor: float = 1.0):
    """Insert a PNG image with fixed width (cm) while preserving aspect ratio."""
    if not image_png:
        return
    buf = BytesIO(image_png)
    img = OpenPyxlImage(buf)

    min_r, min_c, max_r, max_c = _get_merged_box(ws, row, col)

    iw, ih = img.width, img.height
    if not iw or not ih:
        return

    target_w_px = _cm_to_pixels(TARGET_BOM_IMAGE_WIDTH_CM)
    if scale_factor and scale_factor > 0:
        target_w_px = max(1, int(round(target_w_px * scale_factor)))
    scale = target_w_px / iw
    target_w_px = max(1, int(round(iw * scale)))
    target_h_px = max(1, int(round(ih * scale)))

    img.width = target_w_px
    img.height = target_h_px

    # Resize target cell area to match image size (expand only).
    num_cols = max(1, max_c - min_c + 1)
    per_col_px = target_w_px / num_cols
    needed_col_w = _pixels_to_col_width(per_col_px)
    for c in range(min_c, max_c + 1):
        letter = get_column_letter(c)
        cur_w = ws.column_dimensions[letter].width
        cur_w = cur_w if cur_w is not None else 8.43
        ws.column_dimensions[letter].width = max(cur_w, needed_col_w)

    num_rows = max(1, max_r - min_r + 1)
    per_row_px = target_h_px / num_rows
    needed_row_h = _pixels_to_row_height_points(per_row_px)
    for r in range(min_r, max_r + 1):
        cur_h = ws.row_dimensions[r].height
        cur_h = cur_h if cur_h is not None else 15.0
        ws.row_dimensions[r].height = max(cur_h, needed_row_h)

    ws.add_image(img, ws.cell(min_r, min_c).coordinate)
    try:
        if hasattr(img, "anchor") and hasattr(img.anchor, "ext") and img.anchor.ext is not None:
            img.anchor.ext.cx = int(img.width * 9525)
            img.anchor.ext.cy = int(img.height * 9525)
    except Exception:
        pass


# ----------------------------
# PDFì—ì„œ BOM ì´ë¯¸ì§€ ì¶”ì¶œ
# ----------------------------
def extract_graphic_color_cell_images_from_pdf(pdf_path: str) -> Dict[Tuple[str, str, str], bytes]:
    """
    Extract thumbnails inside color columns for the Graphic section.
    Returns mapping: (product, material_name, formatted_color_header) -> PNG bytes
    """
    out: Dict[Tuple[str, str, str], bytes] = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for t in (page.find_tables() or []):
                data = t.extract() or []
                if not data or not data[0]:
                    continue
                header = [clean_text_keep_newlines(x) for x in data[0]]
                header_norm = [normalize_header(x) for x in header]
                if "product" not in header_norm or "materialname" not in header_norm:
                    continue
                
                # "Only for Product Colors" 찾기 (정확 일치 또는 서브스트링)
                idx_only = None
                for ci_h, hn_h in enumerate(header_norm):
                    if hn_h == "onlyforproductcolors":
                        idx_only = ci_h
                        break
                    if "onlyforproductcolors" in hn_h:
                        idx_only = ci_h
                        break
                
                if idx_only is None:
                    continue
                
                idx_product = header_norm.index("product")
                idx_material = header_norm.index("materialname")
                idx_comment = header_norm.index("comment") if "comment" in header_norm else len(header_norm)

                # 병합 여부 체크: CC Number가 포함되면 이 셀부터 컬러
                ofpc_raw = clean_text_keep_newlines(header[idx_only]) if idx_only < len(header) else ""
                has_cc = bool(re.search(r'\b\d{9,}\b', ofpc_raw))
                color_start_idx = idx_only if has_cc else idx_only + 1

                # When merged, clean "Only for Product Colors" prefix from header
                # to match pdf_parser.py's header cleaning (ensures image key consistency)
                if has_cc:
                    cleaned = re.sub(
                        r'(?i)only\s+for\s+product\s+colors?\s*[\n\r]*',
                        '', ofpc_raw
                    ).strip()
                    if cleaned:
                        header[idx_only] = cleaned

                color_cols = list(range(color_start_idx, idx_comment))
                if not color_cols:
                    continue

                current_section = ""
                for r_idx in range(1, len(data)):
                    row = data[r_idx]
                    if not row:
                        continue

                    sec = section_from_cell_text(str(row[0] or ""))
                    if sec:
                        current_section = sec
                        continue

                    if current_section != "Graphic":
                        continue

                    prod = clean_text(row[idx_product] if idx_product < len(row) else "") or "GRAPHIC"
                    material = clean_text(row[idx_material] if idx_material < len(row) else "") or "GRAPHIC"

                    if r_idx >= len(t.rows):
                        continue
                    for ci in color_cols:
                        if ci >= len(t.rows[r_idx].cells):
                            continue
                        bbox = t.rows[r_idx].cells[ci]
                        if not bbox:
                            continue
                        if not _has_embedded_image_in_bbox(page, bbox):
                            continue
                        htxt = format_color_header_text(header[ci] if ci < len(header) else "")
                        if not htxt:
                            continue
                        key = (prod, material, htxt)
                        if key in out:
                            continue
                        try:
                            png_data = _find_fitz_image_for_bbox(pdf_path, page.page_number - 1, bbox)
                            if png_data:
                                out[key] = png_data
                                continue
                            pil = _crop_cell_image(page, bbox, resolution=200)
                            if pil is None:
                                continue
                            buf = BytesIO()
                            pil.save(buf, format="PNG")
                            out[key] = buf.getvalue()
                        except Exception:
                            continue

    _fitz_image_cache.pop((pdf_path, 0), None)  # cleanup hint
    return out


def extract_continuation_graphic_images(
    page,
    table_obj,
    row_to_bomrow_map: Dict[int, int],
    current_block_rows: list,
    header: List[str],
    header_norm: List[str],
    pdf_path: str = "",
) -> Dict[Tuple[str, str, str], bytes]:
    """
    continuation í…Œì´ë¸”ì—ì„œ Graphic í–‰ì˜ ì»¬ëŸ¬ ì´ë¯¸ì§€ë¥¼ ì¶”ì¶œ.
    
    Args:
        page: pdfplumber Page ê°ì²´
        table_obj: pdfplumber Table ê°ì²´ (ì…€ bbox ì •ë³´ í¬í•¨)
        row_to_bomrow_map: {raw_data_idx â†’ BomRow index}
        current_block_rows: í˜„ìž¬ ë¸”ë¡ì˜ BomRow ë¦¬ìŠ¤íŠ¸
        header: ì»¬ëŸ¬ í—¤ë” í…ìŠ¤íŠ¸ ë¦¬ìŠ¤íŠ¸
        header_norm: ì •ê·œí™”ëœ í—¤ë” ë¦¬ìŠ¤íŠ¸
    
    Returns:
        {(product, material_name, formatted_color_header) â†’ PNG bytes}
    """
    out: Dict[Tuple[str, str, str], bytes] = {}

    comment_idx = header_norm.index("comment") if "comment" in header_norm else None
    color_col_indices: List[int] = []
    for ci, hn in enumerate(header_norm):
        if comment_idx is not None and ci == comment_idx:
            continue
        raw = clean_text_keep_newlines(header[ci])
        if raw:
            color_col_indices.append(ci)

    if not color_col_indices:
        return out

    for data_i in range(1, len(table_obj.rows)):
        raw_idx = data_i - 1
        target_i = row_to_bomrow_map.get(raw_idx)
        if target_i is None:
            continue
        if target_i >= len(current_block_rows):
            continue

        brow = current_block_rows[target_i]
        if (brow.category or "").lower() != "graphic":
            continue

        prod = brow.product
        material = brow.material_name

        for ci in color_col_indices:
            if ci >= len(table_obj.rows[data_i].cells):
                continue
            bbox = table_obj.rows[data_i].cells[ci]
            if not bbox:
                continue
            if not _has_embedded_image_in_bbox(page, bbox):
                continue

            htxt = format_color_header_text(header[ci] if ci < len(header) else "")
            if not htxt:
                continue

            key = (prod, material, htxt)
            if key in out:
                continue

            try:
                if pdf_path:
                    png_data = _find_fitz_image_for_bbox(pdf_path, page.page_number - 1, bbox)
                    if png_data:
                        out[key] = png_data
                        continue
                pil = _crop_cell_image(page, bbox, resolution=200)
                if pil is None or _is_blank(pil):
                    continue
                buf = BytesIO()
                pil.save(buf, format="PNG")
                out[key] = buf.getvalue()
            except Exception:
                continue

    return out


def extract_bom_image_map_from_pdf(pdf_path: str) -> Dict[Tuple[str, str, str], bytes]:
    """
    Extract images from BOM Details table 'Image' column for specific sections
    (Packaging and Labels, Graphic).
    Returns mapping: (category, product, material_name) -> PNG bytes
    
    ★ 개선: 연속 테이블(헤더 없는 페이지)도 처리하여 Packaging 이미지 추출
    """
    img_map: Dict[Tuple[str, str, str], bytes] = {}
    current_section: str = ""

    wanted_sections = {"Packaging and Labels", "Graphic"}
    
    # 마지막으로 감지한 헤더 정보 (연속 페이지 처리용)
    last_header_info: Optional[Dict] = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.find_tables() or []
            for t in tables:
                data = t.extract() or []
                if not data or not data[0]:
                    continue

                header = [clean_text_keep_newlines(x) for x in data[0]]
                header_norm = [normalize_header(x) for x in header]

                idx_product = header_norm.index("product") if "product" in header_norm else None
                idx_material = header_norm.index("materialname") if "materialname" in header_norm else None
                idx_image = next(
                    (
                        i
                        for i, hn in enumerate(header_norm)
                        if hn == "image" or ("image" in hn and hn != "designimage")
                    ),
                    None,
                )

                # ── Case 1: 정상 헤더가 있는 테이블 ──
                if idx_product is not None and idx_material is not None and idx_image is not None:
                    
                    last_header_info = {
                        'idx_product': idx_product,
                        'idx_material': idx_material,
                        'idx_image': idx_image,
                    }
                    data_start = 1

                # ── Case 2: 헤더 없는 연속 테이블 (이전 페이지에서 계속) ──
                elif last_header_info is not None:
                    first_cell = clean_text(data[0][0] if data[0] else "")
                    idx_product = last_header_info['idx_product']
                    idx_material = last_header_info['idx_material']
                    idx_image = last_header_info['idx_image']

                    first_row = data[0] if data else []
                    first_material = clean_text(first_row[idx_material] if idx_material < len(first_row) else "")
                    has_required_cols = len(first_row) > max(idx_product, idx_material, idx_image)
                    first_norm = normalize_header(first_cell)

                    # 첫 셀이 숫자/섹션 헤더이거나, 헤더 없이 바로 데이터가 이어지는 경우까지 연속으로 처리
                    is_continuation = (
                        bool(re.fullmatch(r"\d{5,}", first_cell))
                        or section_from_cell_text(first_cell) is not None
                        or (
                            has_required_cols
                            and (first_material or first_cell)
                            and first_norm not in {"product", "materialname"}
                        )
                    )
                    if not is_continuation:
                        continue

                    data_start = 0  # 헤더 행 없이 바로 데이터
                else:
                    continue

                for r_idx in range(data_start, len(data)):
                    row = data[r_idx]
                    if not row:
                        continue

                    row_texts = [clean_text_keep_newlines(x) for x in row]
                    first = row_texts[0] if row_texts else ""
                    sec = section_from_cell_text(first)
                    if sec:
                        current_section = sec
                        continue

                    if current_section not in wanted_sections:
                        continue

                    prod = clean_text(row[idx_product] if idx_product < len(row) else "")
                    material = clean_text(row[idx_material] if idx_material < len(row) else "")
                    if not prod:
                        prod = current_section.upper() if current_section else ""
                    if not prod:
                        continue

                    key = (current_section, prod, material)
                    if key in img_map:
                        continue

                    if r_idx >= len(t.rows):
                        continue
                    if idx_image >= len(t.rows[r_idx].cells):
                        continue
                    bbox = t.rows[r_idx].cells[idx_image]
                    if not bbox:
                        continue

                    try:
                        has_embedded = _has_embedded_image_in_bbox(page, bbox)
                        if has_embedded:
                            png_data = _find_fitz_image_for_bbox(pdf_path, page.page_number - 1, bbox)
                            if png_data:
                                img_map[key] = png_data
                                continue
                        pil = _crop_cell_image(page, bbox, resolution=250)
                        if pil is None:
                            continue
                        if not has_embedded:
                            pil = _trim_pil_to_content(pil)
                        if _is_blank(pil):
                            continue
                        buf = BytesIO()
                        pil.save(buf, format="PNG")
                        img_map[key] = buf.getvalue()
                    except Exception:
                        continue

    return img_map
