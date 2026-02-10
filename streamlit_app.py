"""
BOM PDF -> Excel Auto Filler  (Streamlit Web App)

기존 tkinter GUI의 모든 기능을 100% 동일하게 웹에서 제공합니다.
- 단일 PDF  → 단일 Excel 파일
- 복수 PDF  → 하나의 Excel 파일, PDF별 시트 분리
- 이미지 처리 (Design Image, BOM Row Image, Graphic Color Image) 동일
"""

import os
import sys
import tempfile

import streamlit as st
from openpyxl import load_workbook

# 같은 디렉토리의 모듈을 import 할 수 있도록 경로 보장
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

from excel_writer import fill_template, fill_sheet, sanitize_sheet_name

# ── 페이지 설정 ──────────────────────────────────────────────
st.set_page_config(
    page_title="BOM PDF → Excel 자동 입력",
    page_icon="📋",
    layout="centered",
)

st.title("📋 BOM PDF → Excel 자동 입력")
st.caption("PDF에서 BOM 데이터를 추출하여 Excel 양식에 자동으로 입력합니다.")

DEFAULT_TEMPLATE = os.path.join(_APP_DIR, "양식.xlsx")

# ── Session State 초기화 ─────────────────────────────────────
if "result" not in st.session_state:
    st.session_state.result = None        # (filename, bytes)
if "logs" not in st.session_state:
    st.session_state.logs = []

# ── 1) Excel 양식 선택 ──────────────────────────────────────
st.subheader("1. Excel 양식 선택")

has_default = os.path.exists(DEFAULT_TEMPLATE)
template_options = (
    ["기본 내장 양식 (양식.xlsx)", "직접 업로드"]
    if has_default else ["직접 업로드"]
)
template_option = st.radio("양식을 선택하세요:", options=template_options, horizontal=True)

uploaded_template = None
if template_option == "직접 업로드":
    uploaded_template = st.file_uploader("Excel 양식 파일 (.xlsx)", type=["xlsx"], key="tpl")

# ── 2) BOM PDF 업로드 ───────────────────────────────────────
st.subheader("2. BOM PDF 업로드")
uploaded_pdfs = st.file_uploader(
    "BOM PDF 파일 (복수 선택 가능)",
    type=["pdf"],
    accept_multiple_files=True,
    key="pdfs",
)

# ── 3) 실행 ─────────────────────────────────────────────────
st.subheader("3. 실행")

can_run = bool(uploaded_pdfs)
if template_option == "직접 업로드" and uploaded_template is None:
    can_run = False

if st.button("🚀 실행하기", disabled=not can_run, use_container_width=True, type="primary"):
    st.session_state.result = None
    st.session_state.logs = []
    logs = st.session_state.logs
    total = len(uploaded_pdfs)

    with st.status(f"📋 {total}개 PDF 처리 중...", expanded=True) as status:
        progress = st.progress(0, text="준비 중...")

        with tempfile.TemporaryDirectory() as tmpdir:
            # ── 양식 파일 준비 ──
            if template_option == "직접 업로드":
                tpl_path = os.path.join(tmpdir, "template.xlsx")
                with open(tpl_path, "wb") as f:
                    f.write(uploaded_template.getvalue())
            else:
                tpl_path = DEFAULT_TEMPLATE

            # ── PDF 임시 저장 ──
            pdf_paths = []
            for i, pdf_file in enumerate(uploaded_pdfs):
                p = os.path.join(tmpdir, f"{i}_{pdf_file.name}")
                with open(p, "wb") as f:
                    f.write(pdf_file.getvalue())
                pdf_paths.append(p)

            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            if total == 1:
                # ── 단일 PDF → 단일 파일 ──
                pdf_name = uploaded_pdfs[0].name
                base_name = os.path.splitext(pdf_name)[0]
                out_name = f"{base_name}_filled.xlsx"
                out_path = os.path.join(tmpdir, out_name)

                progress.progress(0, text=f"처리 중: {pdf_name}")
                st.write(f"📄 처리 중: **{pdf_name}**")
                logs.append(f"📄 [1/1] 처리 중: {pdf_name}")

                try:
                    fill_template(tpl_path, pdf_paths[0], out_path)
                    logs.append(f"   ✅ 완료: {out_name}")
                    st.write(f"   ✅ 완료")

                    with open(out_path, "rb") as f:
                        st.session_state.result = (out_name, f.read())
                except Exception as e:
                    logs.append(f"   ❌ 실패: {e}")
                    st.error(f"실패: {e}")

                progress.progress(1.0, text="완료!")

            else:
                # ── 복수 PDF → 하나의 파일, 시트별 분리 ──
                wb = load_workbook(tpl_path)
                original_sheets = list(wb.sheetnames)
                template_ws = wb.active

                sheet_names_used = set()
                success_count = 0
                fail_count = 0

                for idx, (pdf_path, pdf_file) in enumerate(
                    zip(pdf_paths, uploaded_pdfs)
                ):
                    pdf_name = pdf_file.name
                    progress.progress(
                        idx / total,
                        text=f"[{idx + 1}/{total}] {pdf_name}",
                    )
                    st.write(f"📄 [{idx + 1}/{total}] **{pdf_name}**")
                    logs.append(f"📄 [{idx + 1}/{total}] 처리 중: {pdf_name}")

                    try:
                        new_ws = wb.copy_worksheet(template_ws)
                        design_number = fill_sheet(new_ws, pdf_path)

                        # 시트 이름 결정
                        name = design_number or os.path.splitext(pdf_name)[0]
                        name = sanitize_sheet_name(name)
                        base_name = name
                        counter = 1
                        while name in sheet_names_used:
                            suffix = f"_{counter}"
                            name = sanitize_sheet_name(
                                base_name[: 31 - len(suffix)] + suffix
                            )
                            counter += 1
                        sheet_names_used.add(name)
                        new_ws.title = name

                        logs.append(f"   ✅ 완료 → 시트: {name}")
                        st.write(f"   ✅ → 시트: **{name}**")
                        success_count += 1

                    except Exception as e:
                        logs.append(f"   ❌ 실패: {e}")
                        st.write(f"   ❌ 실패: {e}")
                        fail_count += 1

                # 원본 템플릿 시트 모두 삭제
                for sn in original_sheets:
                    if sn in wb.sheetnames:
                        wb.remove(wb[sn])

                out_name = "BOM_combined_filled.xlsx"
                out_path = os.path.join(tmpdir, out_name)
                wb.save(out_path)

                with open(out_path, "rb") as f:
                    st.session_state.result = (out_name, f.read())

                progress.progress(1.0, text="완료!")
                logs.append(
                    f"\n📊 결과: 성공 {success_count}개 / 실패 {fail_count}개"
                )

        status.update(label="✅ 처리 완료!", state="complete")

# ── 4) 결과 다운로드 ────────────────────────────────────────
if st.session_state.result:
    st.subheader("4. 결과 다운로드")
    fname, fbytes = st.session_state.result
    st.download_button(
        label=f"📥 {fname} 다운로드",
        data=fbytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── 5) 처리 로그 ────────────────────────────────────────────
if st.session_state.logs:
    with st.expander("📋 처리 로그", expanded=False):
        st.code("\n".join(st.session_state.logs))
